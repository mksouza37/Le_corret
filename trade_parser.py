import os
import re
import pdfplumber
import pandas as pd
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from PyPDF2 import PdfReader, PdfWriter
import shutil

# === Clean up any old output Excel file ===

if os.path.exists("output_all_invoices.xlsx"):
    os.remove("output_all_invoices.xlsx")

# === Clean temporary directory used for split multi-date PDFs ===
temp_dir = "split_by_date_temp"
if os.path.exists(temp_dir):
    shutil.rmtree(temp_dir)

# === Clear global DataFrames if running interactively (e.g. Colab) ===
import gc
try:
    del df_trades, df_summary, df_consistency
except:
    pass
gc.collect()

# === KEY NORMALIZATION MAP ===
# === RESUMO MAPPING FOR A VISTA ===
RESUMO_KEY_MAP_AVISTA = {
    "Debêntures": ["Debêntures"],
    "Vendas à Vista": ["Vendas à vista", "Venda à Vista", "Vendas à Vista"],
    "Compras à Vista": ["Compras à vista", "Compra à Vista"],
    "Opções - compras": ["Opções - compras", "Compra Opções"],
    "Opções - vendas": ["Opções - vendas", "Venda Opções"],
    "Operações à termo": ["Operações à termo", "Operação a Termo"],
    "Valor das oper. c/ títulos públ. (v. nom.)": [
        "Valor das oper. c/ títulos públ. (v. nom.)",
        "Valor das oper. com títulos públicos",
        "Valor das operações com títulos públicos"
    ],
    "Valor das operações": ["Valor das operações", "Total das operações"],
    "Valor líquido das operações": ["Valor líquido das operações", "Líquido operações"],
    "Taxa de liquidação": ["Taxa de liquidação", "Taxa liquidação"],
    "Taxa de Registro": ["Taxa de Registro", "Registro"],
    "Total CBLC": ["Total CBLC"],
    "Taxa de termo/opções": ["Taxa de termo/opções", "Taxa termo/opções"],
    "Taxa A.N.A.": ["Taxa A.N.A.", "Taxa ANA"],
    "Emolumentos": ["Emolumentos"],
    "Total Bovespa / Soma": ["Total Bovespa / Soma", "Total Bovespa/Soma"],
    "Clearing": ["Clearing", "Taxa Operacional"],
    "Execução": ["Execução"],
    "Execução casa": ["Execução casa", "Taxa de Custódia"],
    "Corretagem": ["Corretagem"],
    "ISS": ["ISS(SÃO PAULO)", "ISS (SÃO PAULO)", "ISS", "ISS* (SÃO PAULO - SP)", "Impostos"],
    "IRRF sobre operações": ["I.R.R.F s/operações", "IRRF s/operações", "I.R.R.F. s/ operações", "IRRF s/ operações"],
    "Outras": ["Outras", "Outros"],
    "Total corretagem / Despesas": ["Total corretagem / Despesas", "Total Custos / Despesas"],
    "Valor a ser Liquidado": ["Valor a ser Liquidado", "Líquido para"]
}

# === RESUMO MAPPING FOR BM&F ===
RESUMO_KEY_MAP_BMF = {
    "Venda disponível": ["Venda disponível"],
    "Compra disponível": ["Compra disponível"],
    "Venda Opções": ["Venda Opções"],
    "Compra Opções": ["Compra Opções"],
    "Valor dos negócios": ["Valor dos negócios", "Vlr de Operação/Ajuste"],
    "IRRF": ["IRRF", "IRRF operacional"],
    "IRRF Day Trade (proj.)": ["IRRF Day Trade (proj.)"],
    "Taxa operacional": ["Taxa operacional"],
    "Taxa registro BM&F": ["Taxa registro BM&F"],
    "Taxas BM&F (emol+f.gar)": ["Taxas BM&F (emol+f.gar)"],
    "Outros Custos": ["Outros Custos"],
    "ISS": ["ISS", "Impostos"],
    "Ajuste de posição": ["Ajuste de posição"],
    "Ajuste day trade": ["Ajuste day trade"],
    "Total das despesas": ["Total das despesas", "Total de custos operacionais"],
    "Outros": ["Outros"],
    "IRRF Corretagem": ["IRRF Corretagem"],
    "Total Conta Investimento": ["Total Conta Investimento"],
    "Total Conta Normal": ["Total Conta Normal"],
    "Total líquido (#)": ["Total líquido (#)"],
    "Total líquido da nota": ["Total líquido da nota"],
    "Valor das operações": ["Valor dos negócios", "Vlr de Operação/Ajuste"]  # For consistency check
}

import unicodedata

def remove_accents(text):
    return ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')

def classify_invoice_type(text: str) -> str:
    normalized = remove_accents(text.lower())

    # Keep original A VISTA logic
    if any(x in normalized for x in [
        "negocios realizados", "resumo dos negocios", "negocios efetuados"
    ]) and not any(term in normalized for term in [
            "Mercadoria", "c/v mercadoria"
    ]):
        return "avista"

    # Keep original BM&F logic
    if "c/v mercadoria vencimento" in normalized and "ajuste de posicao" in normalized:
        return "bmf"

    # ✅ Add XP-specific fallback
    #if "corretora xp" in normalized:
    if any(term in normalized for term in [
            "ajuste de posicao", "ajuste day trade", "taxa registro bmf",
            "irrf day trade", "vencimento", "c/v mercadoria", "Mercadoria"
        ]):
            return "bmf"

    return "unknown"


# === MULTI-DATE PDF HANDLING HELPERS ===

def extract_date_and_type_per_page(file_path):
    results = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            # Extract date
            found_date = None
            for pattern in [
                r"Data\s+Preg[aã]o\s*(?:\n|\r|\s)*(\d{2}/\d{2}/\d{4})",
                r"(\d{2}/\d{2}/\d{4})"
            ]:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    found_date = match.group(1)
                    break

            # Extract type
            tipo = classify_invoice_type(text)
            results.append((found_date, tipo))
    return results


def group_pages_by_date_and_type(date_type_pairs):
    groups = []
    current_key = None
    current_pages = []

    for i, (date, tipo) in enumerate(date_type_pairs):
        if date is None or tipo == "unknown":
            continue
        key = (date, tipo)
        if key != current_key:
            if current_pages:
                groups.append((current_key, current_pages))
            current_key = key
            current_pages = [i]
        else:
            current_pages.append(i)
    if current_pages:
        groups.append((current_key, current_pages))
    return groups

def prepare_files_for_processing(pdf_files):
    output_dir = "split_by_date_temp"
    os.makedirs(output_dir, exist_ok=True)
    files_to_process = []

    for file_path in pdf_files:
        date_type_pairs = extract_date_and_type_per_page(file_path)

        unique_keys = list(set(k for k in date_type_pairs if k[0] and k[1] != "unknown"))

        if len(unique_keys) <= 1:
            files_to_process.append(file_path)
        else:
            groups = group_pages_by_date_and_type(date_type_pairs)
            reader = PdfReader(file_path)
            for (date_str, tipo), page_numbers in groups:
                writer = PdfWriter()
                for page_num in page_numbers:
                    writer.add_page(reader.pages[page_num])
                filename_base = os.path.splitext(os.path.basename(file_path))[0]
                tipo_str = tipo.upper().replace(" ", "")
                new_filename = f"{filename_base}_{date_str.replace('/', '-')}_{tipo_str}.pdf"
                out_path = os.path.join(output_dir, new_filename)
                with open(out_path, "wb") as f:
                    writer.write(f)
                files_to_process.append(out_path)

    return files_to_process

@dataclass
class BrokerConfig:
    name: str
    invoice_patterns: List[str]
    date_patterns: List[str]
    client_patterns: Dict[str, str]
    trade_start_marker: str
    trade_end_marker: str
    signature_patterns: List[str]

BTG_CONFIG = BrokerConfig(
    name="BTG",
    invoice_patterns=[
        r"Nota\s+de\s+Negociação\s+N(?:º|o|\u00b0)?\s*[:\-]?\s*(\d+)",
        r"Nr\.?\s*nota\s*[:\-]?\s*(\d+)",
        r"Nota\s*[:\-]?\s*(\d+)"
    ],
    date_patterns=[
        r'Data\s+preg[aã]o\s*(?:\n|\r|\s)*(\d{2}/\d{2}/\d{4})',
        r'(\d{2}/\d{2}/\d{4})'
    ],
    client_patterns={},
    trade_start_marker="Negócios realizados",
    trade_end_marker="Resumo dos Negócios",
    signature_patterns=[r"BTG\s+Pactual", r"BTG\s+Corretora"]
)

ITAU_CONFIG = BrokerConfig(
    name="ITAU",
    invoice_patterns=[
        r"Nr\.?\s*Nota\s*(?:Folha)?\s*(?:\d+\s+)?(\d+)",
    ],
    date_patterns=[
        r"Data\s+Preg[aã]o\s*(?:\n|\r|\s)*(\d{2}/\d{2}/\d{4})",
        r'(\d{2}/\d{2}/\d{4})'
    ],
    client_patterns={},
    trade_start_marker="Negócios Realizados",
    trade_end_marker="Resumo de negócios",
    signature_patterns=[r"Ita[úu]\s+Corretora", r"ITA[ÚU] UNIBANCO"],
)

AGORA_CONFIG = BrokerConfig(
    name="AGORA",
    invoice_patterns=[
        r"Nota\s+de\s+Corretagem\s*Nr\.?\s*Nota\s*(?:Folha)?\s*(?:\d+\s+)?(\d+)",
        r"Nr\.?\s*Nota\s*(?:\d+\s+)?(\d+)"
    ],
    date_patterns=[
        r"Data\s+preg[aã]o\s*(?:\n|\r|\s)*(\d{2}/\d{2}/\d{4})",
        r'(\d{2}/\d{2}/\d{4})'
    ],
    client_patterns={},
    trade_start_marker="Negocios Realizados",
    trade_end_marker="Resumo dos Negócios",
    signature_patterns=[
        r"AGORA\s+CORRETORA", r"agorainvestimentos\.com\.br"],
)

XP_CONFIG = BrokerConfig(
    name="XP",
    invoice_patterns=[
        r"NOTA\s+DE\s+NEGOCIAÇÃO\s+Nr\.?\s*nota\s*(\d{3}[.,]?\d{3})",  # ← Precise
        r"Nr\.?\s*nota\s*(\d{3}[.,]?\d{3})"
    ],
    date_patterns=[
        r"Data\s+preg[aã]o\s*(?:\n|\r|\s)*(\d{2}/\d{2}/\d{4})",
        r'(\d{2}/\d{2}/\d{4})'
    ],
    client_patterns={},
    trade_start_marker="Negócios realizados",
    trade_end_marker="Resumo dos Negócios",
    signature_patterns=[
        r"XP\s+INVESTIMENTOS\s+CORRETORA", r"xpi\.com\.br"]
)

# === PARSER IMPLEMENTATION ===

class GenericParser:
    def __init__(self, config: BrokerConfig):
        self.config = config

    def parse_pdf(self, file_path: str) -> Dict:
        text = self._extract_text(file_path)
        top_fields = self._extract_top_table_fields(text)
        cpf = self._extract_top_client_fields(text)

        invoice = top_fields.get("invoice")
        if not invoice:
            invoice = self._extract_first_match(text, self.config.invoice_patterns)

        # Normalize XP-style dotted invoice numbers like 733.612 → 733612
        if invoice and isinstance(invoice, str):
          invoice = re.sub(r"[^\d]", "", invoice)

        return {
            "broker": self.config.name,
            "invoice": invoice,
            "date": self._extract_first_match(text, self.config.date_patterns),
            "client_cpf": cpf,
            "trades": self._extract_trades(text),
            "summary": self._extract_summary_values(text)
        }

    def _extract_text(self, file_path: str) -> str:
        with pdfplumber.open(file_path) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)

    def _extract_top_client_fields(self, text: str) -> str:
        lines = text.splitlines()
        for i, line in enumerate(lines):
            if "cpf" in line.lower():
                cpf_match = re.search(r'(\d{3}[\.\s]?\d{3}[\.\s]?\d{3}[-\s]?\d{2})', line)
                if not cpf_match and i + 1 < len(lines):
                    cpf_match = re.search(r'(\d{3}[\.\s]?\d{3}[\.\s]?\d{3}[-\s]?\d{2})', lines[i + 1])
                if cpf_match:
                    return cpf_match.group(1).replace(" ", "").strip()
        return ""

    def _extract_top_table_fields(self, text: str) -> Dict[str, str]:
        lines = text.splitlines()
        info = {}

        for i, line in enumerate(lines):
            normalized = line.lower()

            if "nr.nota" in normalized and "data pregão" in normalized:
                if i + 1 < len(lines):
                    next_line = lines[i + 1]
                    matches = re.findall(r'\d{2}/\d{2}/\d{4}|\d+', next_line)
                    if matches:
                        info["invoice"] = matches[0]
                        break

            match_inline = re.search(r"nota\s*(?:nº|no|n°|num)?[\s:]*\s*(\d{4,})", line, re.IGNORECASE)
            if match_inline:
                info["invoice"] = match_inline.group(1)
                break

            if any(k in normalized for k in ["nr. nota", "nr nota", "nº nota", "nota de negociação"]):
                if i + 1 < len(lines):
                    nextline_match = re.search(r'\d{4,}', lines[i + 1])
                    if nextline_match:
                        info["invoice"] = nextline_match.group(0)
                        break

        return info

    def _extract_first_match(self, text: str, patterns: List[str]) -> str:
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL | re.MULTILINE)
            if match:
                return match.group(1).strip()
        return ""

    def _extract_trades(self, text: str) -> List[Dict]:
        trades = []
        parser_name = self.config.name.upper()

        # Extract all trade blocks
        trade_blocks = re.findall(
            rf"{self.config.trade_start_marker}.*?(?={self.config.trade_end_marker}|$)",
            text,
            flags=re.DOTALL | re.IGNORECASE,
        )

        for block in trade_blocks:
            lines = block.splitlines()
            header_idx = None

            expected_labels = ["negociação", "c/v", "tipo", "quantidade", "preço", "valor"]
            for i in range(len(lines) - 1):
                chunk = ' '.join(lines[i:i + 2]).lower()
                if sum(lbl in chunk for lbl in expected_labels) >= 4:
                    header_idx = i + 2
                    break

            if header_idx is None:
                continue

            for line in lines[header_idx:]:
                if not line.strip() or "resumo" in line.lower():
                    break
                try:
                    tokens = line.strip().split()
                    if "XP" in parser_name or "BTG" in parser_name or "ITAU" in parser_name:
                        trade = {
                            "Negociação": tokens[0],
                            "C/V": tokens[1],
                            "Tipo Mercado": tokens[2],
                            "Especificação do Título": " ".join(tokens[3:5]),
                            "Quantidade": int(tokens[-4].replace('.', '').replace(',', '')),
                            "Preço / Ajuste": self._clean_numeric(tokens[-3]),
                            "Valor Operação / Ajuste": self._clean_numeric(tokens[-2]),
                            "D/C": tokens[-1]
                        }

                    elif "AGORA" in parser_name:
                        trade = {
                            "Negociação": ' '.join(tokens[0:2]),
                            "C/V": tokens[3],
                            "Tipo Mercado": tokens[4],
                            "Especificação do Título": ' '.join(tokens[5:7]),
                            "Quantidade": int(tokens[-4].replace('.', '').replace(',', '')),
                            "Preço / Ajuste": self._clean_numeric(tokens[-3]),
                            "Valor Operação / Ajuste": self._clean_numeric(tokens[-2]),
                            "D/C": tokens[-1]
                        }

                    else:
                        continue

                    trade["Tipo"] = "A Vista"  # ✅ Add this line
                    trades.append(trade)

                except Exception as e:
                    print(f"⚠️ Error processing line: {line}\n{e}")
                    continue

        return trades

    def _extract_summary_values(self, text: str) -> Dict[str, float]:
        summary = {}

        # 1. Try to extract the last matching summary block
        matches = list(re.finditer(r"(Resumo dos Negócios.*?)(?=Resumo dos Negócios|$)", text, flags=re.DOTALL | re.IGNORECASE))
        if not matches:
            matches = list(re.finditer(r"(Resumo de negócios.*?)(?=Resumo de negócios|$)", text, flags=re.DOTALL | re.IGNORECASE))
        if not matches:
            return {k: 0.0 for k in RESUMO_KEY_MAP_AVISTA}  # return zeros if nothing found

        last_summary_text = matches[-1].group(1)
        joined = re.sub(r"\s{2,}", " ", " ".join(last_summary_text.splitlines()))

        key_map = RESUMO_KEY_MAP_AVISTA if isinstance(self, AVistaParser) else RESUMO_KEY_MAP_BMF
        for std_key, variants in key_map.items():
            found = False
            for var in variants:
                if std_key == "Valor a ser Liquidado":
                    match = re.search(
                        r"Líquido para\s+\d{2}/\d{2}/\d{4}(?:\s+\d{2}:\d{2}:\d{2})?\s*([\d\.,]+)\s*([CD])?",
                        joined,
                        flags=re.IGNORECASE
                    )
                    if match:
                        value = abs(self._clean_numeric(match.group(1)))
                        label = match.group(2) or ""
                        if value == 0:
                            label = ""
                        summary[std_key] = value
                        summary[f"{std_key} AAAA"] = label
                        found = True
                        break

                elif std_key == "IRRF sobre operações":
                    # Try to find IRRF above the "Outras" line
                    summary_lines = last_summary_text.splitlines()
                    outras_variants = RESUMO_KEY_MAP_AVISTA["Outras"]
                    found_line_index = -1

                    for i, line in enumerate(summary_lines):
                        if any(alias.lower() in line.lower() for alias in outras_variants):
                            found_line_index = i
                            break

                    if found_line_index > 0:
                        previous_line = summary_lines[found_line_index - 1]
                        matches = re.findall(r"([\d\.,]+)\s*([CD])?", previous_line)
                        if matches:
                            number_str, label = matches[-1]
                            value = abs(self._clean_numeric(number_str))
                            label = label or ""
                            if value == 0:
                                label = ""
                            summary[std_key] = value
                            summary[f"{std_key} AAAA"] = label
                            found = True
                            break

                else:
                    # Generic regex match
                    pattern = fr"{re.escape(var)}\s*[.:\-]*\s*R?\$?\s*([\d\.,]+)\s*([CD])?"
                    match = re.search(pattern, joined, flags=re.IGNORECASE)

                    if match:
                        value = abs(self._clean_numeric(match.group(1)))
                        label = match.group(2) or ""
                        if value == 0:
                            label = ""
                        summary[std_key] = value
                        summary[f"{std_key} AAAA"] = label
                        found = True
                        break

            if not found:
                summary[std_key] = 0.0
                summary[f"{std_key} AAAA"] = ""

        # Special total computation for ITAU
        if self.config.name.upper() == "ITAU":
            summary["Total corretagem / Despesas"] = sum(
                summary.get(key, 0.0) for key in [
                    "Corretagem", "ISS", "IRRF sobre operações", "Outras"
                ]
            )

        return summary

    def _clean_numeric(self, value: str) -> float:
        try:
            return float(value.replace('.', '').replace(',', '.'))
        except:
            return 0.0

class AVistaParser(GenericParser):

    pass  # Inherits the default behavior

class BMFParser(GenericParser):

    def _extract_trades(self, text: str) -> List[Dict]:
        trades = []
        lines = text.splitlines()
        for line in lines:
            tokens = line.strip().split()
            if len(tokens) >= 9 and re.match(r"^[CV]$", tokens[0]) and re.match(r"\d{2}/\d{2}/\d{4}", tokens[2]):
                try:
                    trade = {
                        "Tipo": "BM&F",
                        "C/V": tokens[0],
                        "Mercadoria": tokens[1],
                        "Vencimento": tokens[2],
                        "Quantidade": int(tokens[3]),
                        "Preço / Ajuste": self._clean_numeric(tokens[4]),
                        "Tipo Negócio": tokens[5],
                        "Valor Operação": self._clean_numeric(tokens[6]),
                        "D/C": tokens[7],
                        "Taxa Operacional": self._clean_numeric(tokens[8])
                    }
                    trades.append(trade)
                except Exception as e:
                    print(f"⚠️ BM&F trade parse error: {line}\n{e}")
        return trades

    def parse_pdf(self, file_path: str) -> Dict:
        text = self._extract_text(file_path)
        top_fields = self._extract_top_table_fields(text)
        cpf = self._extract_top_client_fields(text)

        invoice = top_fields.get("invoice")
        if not invoice:
            invoice = self._extract_first_match(text, self.config.invoice_patterns)

        return {
            "broker": self.config.name,
            "invoice": invoice,
            "date": self._extract_first_match(text, self.config.date_patterns),
            "client_cpf": cpf,
            "trades": self._extract_trades(text),
            "summary": self._extract_summary_values(text, file_path)
        }

# === BROKER-SPECIFIC PARSERS ===

# --- XP ---
class XPAvistaParser(AVistaParser):
    pass  # Inherits default AVista behavior for now

class XPBMFParser(BMFParser):

    def parse_pdf(self, file_path: str) -> Dict:
      text = self._extract_text(file_path)
      top_fields = self._extract_top_table_fields(text)
      cpf = self._extract_top_client_fields(text)

      invoice = self._extract_invoice_number(text)
      date = self._extract_first_match(text, self.config.date_patterns)

      return {
          "broker": self.config.name,
          "invoice": invoice,
          "date": date,
          "client_cpf": cpf,
          "trades": self._extract_trades(text),
          "summary": self._extract_summary_values(text, file_path)
      }

    def _extract_invoice_number(self, text: str) -> str:
        # Look for pattern: "NOTA DE NEGOCIAÇÃO Nr. nota" followed by the number
        lines = text.splitlines()
        for i, line in enumerate(lines):
            if re.search(r"NOTA\s+DE\s+NEGOCIAÇÃO\s+Nr\.?\s*nota", line, re.IGNORECASE):
                # Try next line
                if i + 1 < len(lines):
                    next_line = lines[i + 1]
                    match = re.search(r"(\d{3}[.,]?\d{3})", next_line)
                    if match:
                        return re.sub(r"[^\d]", "", match.group(1))  # Normalize "733.612" → "733612"
        # Fallback if not found
        return self._extract_first_match(text, self.config.invoice_patterns)

    def _extract_trades(self, text: str) -> List[Dict]:
        trades = []
        lines = text.splitlines()

        for line in lines:
            tokens = line.strip().split()
            if len(tokens) < 10:
                continue

            if (
                re.match(r"^[CV]$", tokens[0]) and
                re.search(r"\d{2}/\d{2}/\d{4}", tokens[3])
            ):
                try:
                    mercadoria = f"{tokens[1]} {tokens[2]}"
                    vencimento = tokens[3].replace('@', '')  # Fix for @18/08/2021
                    quantidade = int(tokens[4].replace('.', ''))
                    preco_ajuste = self._clean_numeric(tokens[5])

                    # Handle possible split "DAY TRADE"
                    if tokens[6] == "DAY" and tokens[7] == "TRADE":
                        tipo_negocio = "DAY TRADE"
                        valor_operacao = self._clean_numeric(tokens[8])
                        dc = tokens[9]
                        taxa_operacional = self._clean_numeric(tokens[10])
                    else:
                        tipo_negocio = tokens[6]
                        valor_operacao = self._clean_numeric(tokens[7])
                        dc = tokens[8]
                        taxa_operacional = self._clean_numeric(tokens[9])

                    trade = {
                        "Tipo": "BM&F",
                        "C/V": tokens[0],
                        "Mercadoria": mercadoria,
                        "Vencimento": vencimento,
                        "Quantidade": quantidade,
                        "Preço / Ajuste": preco_ajuste,
                        "Tipo Negócio": tipo_negocio,
                        "Valor Operação": valor_operacao,
                        "D/C": dc,
                        "Taxa Operacional": taxa_operacional
                    }
                    trades.append(trade)

                except Exception as e:
                    print(f"⚠️ XP BM&F trade parse error: {line}\n{e}")
                    continue

        return trades



    def _extract_summary_values(self, text: str, file_path: str) -> Dict[str, float]:
        summary = {}

        LABEL_POSITIONS = {
            "Venda disponível": (0, 0),
            "Compra disponível": (0, 1),
            "Venda Opções": (0, 2),
            "Compra Opções": (0, 3),
            "Valor dos negócios": (0, 4),
            "IRRF": (1, 0),
            "IRRF Day Trade (proj.)": (1, 1),
            "Taxa operacional": (1, 2),
            "Taxa registro BM&F": (1, 3),
            "Taxas BM&F (emol+f.gar)": (1, 4),
            "Outros Custos": (2, 0),
            "ISS": (2, 1),
            "Ajuste de posição": (2, 2),
            "Ajuste day trade": (2, 3),
            "Total das despesas": (2, 4),
            "Outros": (3, 0),
            "IRRF Corretagem": (3, 1),
            "Total Conta Investimento": (3, 2),
            "Total Conta Normal": (3, 3),
            "Total líquido (#)": (3, 4),
            "Total líquido da nota": (3, 5),
            "Valor das operações": (0, 4)  # mirror from "Valor dos negócios"
        }

        for key in LABEL_POSITIONS:
            summary[key] = 0.0
            summary[f"{key} AAAA"] = ""

        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    lines = [line.strip() for line in page_text.splitlines() if line.strip()]

                    label_blocks = [[] for _ in range(4)]
                    number_blocks = [[] for _ in range(4)]
                    block_index = -1

                    for line in lines:
                        if any(label in line for label in LABEL_POSITIONS):
                            block_index += 1
                            label_blocks[block_index] = re.split(r"\s{2,}", line)
                        elif re.search(r"\d", line) and block_index >= 0:
                            clean = re.sub(r"[^\d.,\sD]", "", line)
                            number_blocks[block_index] += [val.strip() for val in clean.split() if re.search(r"[\d,]", val)]

                    # Special cleaning for block 3
                    if len(number_blocks) > 3:
                        combined = " ".join(number_blocks[3])
                        number_blocks[3] = re.findall(r"[\d.,]+", combined)

                    for label, (block_idx, pos_idx) in LABEL_POSITIONS.items():
                        try:
                            value_str = number_blocks[block_idx][pos_idx]
                            value = self._clean_numeric(value_str)
                            letter = "C"

                            for line in lines:
                                if value_str in line:
                                    after = line.split(value_str, 1)[1]
                                    if re.search(r'\bD\b', after):
                                        letter = "D"
                                    break

                            summary[label] = abs(value)
                            summary[f"{label} AAAA"] = letter if abs(value) > 0 else ""

                        except (IndexError, ValueError):
                            summary[label] = 0.0
                            summary[f"{label} AAAA"] = ""

        except Exception as e:
            print(f"❌ Failed to extract BM&F summary with positional logic: {e}")
            for label in LABEL_POSITIONS:
                summary[label] = 0.0
                summary[f"{label} AAAA"] = ""

        self.last_summary_columns = list(summary.keys())  # ✅ Needed for D/C column control
        return summary

# --- BTG ---
class BTGAvistaParser(AVistaParser):
    pass  # Inherits default AVista behavior for now

class BTGBMFParser(BMFParser):

    def _extract_summary_values(self, text: str, file_path: str) -> Dict[str, float]:
        summary = {}

        LABEL_POSITIONS = {
            "Venda disponível": (0, 0),
            "Compra disponível": (0, 1),
            "Venda Opções": (0, 2),
            "Compra Opções": (0, 3),
            "Valor dos negócios": (0, 4),
            "IRRF": (1, 0),
            "IRRF Day Trade (proj.)": (1, 1),
            "Taxa operacional": (1, 2),
            "Taxa registro BM&F": (1, 3),
            "Taxas BM&F (emol+f.gar)": (1, 4),
            "Outros Custos": (2, 0),
            "ISS": (2, 1),
            "Ajuste de posição": (2, 2),
            "Ajuste day trade": (2, 3),
            "Total das despesas": (2, 4),
            "Outros": (3, 0),
            "IRRF Corretagem": (3, 1),
            "Total Conta Investimento": (3, 2),
            "Total Conta Normal": (3, 3),
            "Total líquido (#)": (3, 4),
            "Total líquido da nota": (3, 5),
            "Valor das operações": (0, 4)  # mirror from "Valor dos negócios"
        }

        for key in LABEL_POSITIONS:
            summary[key] = 0.0
            summary[f"{key} AAAA"] = ""

        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    lines = [line.strip() for line in page_text.splitlines() if line.strip()]

                    label_blocks = [[] for _ in range(4)]
                    number_blocks = [[] for _ in range(4)]
                    block_index = -1

                    for line in lines:
                        if any(label in line for label in LABEL_POSITIONS):
                            block_index += 1
                            label_blocks[block_index] = re.split(r"\s{2,}", line)
                        elif re.search(r"\d", line) and block_index >= 0:
                            clean = re.sub(r"[^\d.,\sD]", "", line)
                            number_blocks[block_index] += [val.strip() for val in clean.split() if re.search(r"[\d,]", val)]

                    # Special cleaning for block 3
                    if len(number_blocks) > 3:
                        combined = " ".join(number_blocks[3])
                        number_blocks[3] = re.findall(r"[\d.,]+", combined)

                    for label, (block_idx, pos_idx) in LABEL_POSITIONS.items():
                        try:
                            value_str = number_blocks[block_idx][pos_idx]
                            value = self._clean_numeric(value_str)
                            letter = "C"

                            for line in lines:
                                if value_str in line:
                                    after = line.split(value_str, 1)[1]
                                    if re.search(r'\bD\b', after):
                                        letter = "D"
                                    break

                            summary[label] = abs(value)
                            summary[f"{label} AAAA"] = letter if abs(value) > 0 else ""

                        except (IndexError, ValueError):
                            summary[label] = 0.0
                            summary[f"{label} AAAA"] = ""

        except Exception as e:
            print(f"❌ Failed to extract BM&F summary with positional logic: {e}")
            for label in LABEL_POSITIONS:
                summary[label] = 0.0
                summary[f"{label} AAAA"] = ""

        self.last_summary_columns = list(summary.keys())  # ✅ Needed for D/C column control
        return summary

# === PROCESS MULTIPLE FILES ===

class TradeProcessor:
    PARSERS = [
        GenericParser(BTG_CONFIG),
        GenericParser(ITAU_CONFIG),
        GenericParser(AGORA_CONFIG),
        GenericParser(XP_CONFIG),
    ]

    BROKER_TYPE_TO_PARSER = {
        ("XP", "avista"): XPAvistaParser,
        ("XP", "bmf"): XPBMFParser,
        ("BTG", "avista"): BTGAvistaParser,
        ("BTG", "bmf"): BTGBMFParser,
        ("ITAU", "avista"): AVistaParser,
        ("AGORA", "avista"): AVistaParser,
        # Add more as needed
    }

    @classmethod
    def process_pdfs(cls, file_paths: List[str]):
        all_trades, all_summaries = [], []

        for file_path in file_paths:
            try:
                with pdfplumber.open(file_path) as pdf:
                    text = "\n".join(page.extract_text() or "" for page in pdf.pages)

                matched_parser = None
                for parser in cls.PARSERS:
                    if cls._match_broker_by_signature(text, parser):
                        invoice_type = classify_invoice_type(text)
                        broker_name = parser.config.name.upper()

                        parser_class = cls.BROKER_TYPE_TO_PARSER.get(
                            (broker_name, invoice_type),
                            AVistaParser if invoice_type == "avista" else BMFParser  # fallback default
                        )

                        matched_parser = parser_class(parser.config)

                        print(f"🔍 Using {parser_class.__name__} for broker '{broker_name}' and type '{invoice_type}'")
                        break

                if not matched_parser:
                    print(f"⚠️ No matching parser found for {file_path}")
                    continue

                result = matched_parser.parse_pdf(file_path)

                trades = result.get("trades", [])
                summary = result.get("summary", {})
                invoice = result.get("invoice", "")

                for trade in trades:
                    trade.update({
                        "broker": result["broker"],
                        "date": result["date"],
                        "invoice": invoice,
                        "client_cpf": result.get("client_cpf", "")
                    })
                    all_trades.append(trade)

                if summary:
                  summary_row = {
                      "invoice": invoice,
                      "broker": result["broker"],
                      "Tipo": trade.get("Tipo", "Unknown"), # Add 'Tipo' to summary row
                      **summary
                  }
                  all_summaries.append(summary_row)

            except Exception as e:
                print(f"❌ Error processing {file_path}: {e}")

        return all_trades, all_summaries

    @staticmethod
    def _match_broker_by_signature(text: str, parser: GenericParser) -> bool:
        for pattern in parser.config.signature_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return True
        return False

    @classmethod
    def process_directory(cls, directory: str) -> pd.DataFrame:
        pdf_files = [os.path.join(directory, f) for f in os.listdir(directory) if f.lower().endswith(".pdf")]
        if not pdf_files:
            print("❌ No PDF files found in directory.")
            return pd.DataFrame()

        prepared_files = prepare_files_for_processing(pdf_files)
        trades, summaries = TradeProcessor.process_pdfs(prepared_files)
        df_trades = pd.DataFrame(trades)

        # === Valor de referência principal para negócios individuais ===
        if "Valor Operação" in df_trades.columns and "Valor Operação / Ajuste" in df_trades.columns:
            df_trades["valor_trades"] = df_trades.apply(
                lambda row: row["Valor Operação"] if str(row.get("Tipo", "")).strip().lower() == "bm&f" else row[
                    "Valor Operação / Ajuste"],
                axis=1
            )
        elif "Valor Operação" in df_trades.columns:
            df_trades["valor_trades"] = df_trades["Valor Operação"]
        elif "Valor Operação / Ajuste" in df_trades.columns:
            df_trades["valor_trades"] = df_trades["Valor Operação / Ajuste"]
        else:
            df_trades["valor_trades"] = 0

        # === Consistência: ajusta BM&F com D/C == C para valor negativo
        df_trades["valor_trades_consistency"] = df_trades["valor_trades"]
        if not df_trades.empty and "Tipo" in df_trades.columns and "D/C" in df_trades.columns:
            bmf_mask = df_trades["Tipo"].str.lower().str.strip() == "bm&f"
            c_mask = df_trades["D/C"].str.upper().str.strip() == "C"
            df_trades.loc[bmf_mask & c_mask, "valor_trades_consistency"] *= -1

        df_summary = pd.DataFrame(summaries).dropna(how='all')

        # === Limpa colunas AAAA que estão totalmente vazias
        if not df_summary.empty:
            for col in df_summary.columns:
                if col.endswith("AAAA") and df_summary[col].replace("", pd.NA).isna().all():
                    df_summary.drop(columns=[col], inplace=True)

        # === CONSISTÊNCIA ENTRE NEGÓCIOS E RESUMO ===
        if "invoice" in df_trades.columns or "invoice" in df_summary.columns:

            #trade_value_column = "valor_trades"
            trade_value_column = "valor_trades_consistency"


            # 1. Totais de negócios com valor ajustado
            trade_totals = (
                df_trades.groupby(["invoice", "broker", "Tipo"])["valor_trades_consistency"]
                .sum()
                .reset_index()
                .rename(columns={"valor_trades_consistency": trade_value_column})
            )
            print(trade_totals)

            # 2. Valor total do resumo (usa diferentes possíveis colunas)
            summary_candidates = ["Valor das operações", "Valor dos negócios"]
            valor_col = None
            for col in summary_candidates:
                if col in df_summary.columns:
                    valor_col = col
                    break

            if valor_col:
                resumo_valores = df_summary[["invoice", "broker", "Tipo", valor_col]].rename(
                    columns={valor_col: "valor_das_operacoes"}
                )

                all_invoices = pd.DataFrame(
                    pd.concat([
                        trade_totals[["invoice", "broker", "Tipo"]],
                        resumo_valores[["invoice", "broker", "Tipo"]]
                    ]).drop_duplicates(), columns=["invoice", "broker", "Tipo"]
                )

                df_consistency = all_invoices \
                    .merge(trade_totals, on=["invoice", "broker", "Tipo"], how="left") \
                    .merge(resumo_valores, on=["invoice", "broker", "Tipo"], how="left")

                df_consistency[trade_value_column] = df_consistency[trade_value_column].fillna(0)
                df_consistency["valor_das_operacoes"] = df_consistency["valor_das_operacoes"].fillna(0)
                df_consistency["Diferença"] = (
                    df_consistency[trade_value_column] - df_consistency["valor_das_operacoes"]
                ).round(2)
                df_consistency["Status"] = df_consistency["Diferença"].apply(
                    lambda x: "OK" if abs(x) < 0.01 else "Inconsistência"
                )

            else:
                df_consistency = pd.DataFrame()
        else:
            df_consistency = pd.DataFrame()

        # === RENOMEIA COLUNAS PARA EXPORTAÇÃO ===
        df_trades.rename(columns={
            "invoice": "Número da Nota", "broker": "Corretora",
            "client_cpf": "CPF", "date": "Data da Operação"
        }, inplace=True)

        df_summary.rename(columns={
            "invoice": "Número da Nota", "broker": "Corretora"
        }, inplace=True)

        if not df_consistency.empty:
            df_consistency.rename(columns={
                "invoice": "Número da Nota", "broker": "Corretora"
            }, inplace=True)

        # === CPF para nome de arquivo
        cpf_value = (
            df_trades["CPF"].iloc[0].replace('.', '').replace('-', '')
            if "CPF" in df_trades.columns and not df_trades.empty
            else "unknown"
        )
        output_file = os.path.join("tmp", f"trades_output - {cpf_value}.xlsx")

        def autofit_columns(worksheet):
            from openpyxl.utils import get_column_letter

            column_widths = {}

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value:
                        col_letter = get_column_letter(cell.column)
                        current_width = column_widths.get(col_letter, 0)
                        value_length = len(str(cell.value))
                        column_widths[col_letter] = max(current_width, value_length)

                        # Center align headers (bold cells)
                        if cell.font and cell.font.bold:
                            cell.alignment = Alignment(horizontal="center")

            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width + 2  # +2 for padding

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            workbook = writer.book

            # === Negócios ===
            ws = workbook.create_sheet(title="Negócios")
            row_idx = 1

            if not df_trades.empty:
                tipos_ordenados = ["A Vista", "BM&F"]
                for tipo in tipos_ordenados:
                    tipo_lower = tipo.strip().lower()
                    if tipo_lower not in df_trades["Tipo"].str.lower().str.strip().unique():
                        continue

                    block = df_trades[df_trades["Tipo"].str.lower().str.strip() == tipo_lower].copy()
                    block = block.sort_values(by=["Corretora", "Data da Operação"])

                    if tipo_lower == "a vista":
                        block_columns = [
                            "CPF", "Tipo", "Corretora", "Data da Operação", "Número da Nota",
                            "Negociação", "C/V", "Tipo Mercado", "Especificação do Título",
                            "Quantidade", "Preço / Ajuste", "Valor Operação / Ajuste", "D/C"
                        ]
                    elif tipo_lower == "bm&f":
                        block_columns = [
                            "CPF", "Tipo", "Corretora", "Data da Operação", "Número da Nota",
                            "C/V", "Mercadoria", "Vencimento", "Quantidade", "Preço / Ajuste",
                            "Tipo Negócio", "Valor Operação", "D/C", "Taxa Operacional"
                        ]
                    else:
                        block_columns = block.columns.tolist()

                    block = block[[col for col in block_columns if col in block.columns]]

                    ws.cell(row=row_idx, column=1, value=f"*** {tipo.upper()} ***").font = Font(bold=True)

                    row_idx += 2

                    for col_idx, col_name in enumerate(block.columns, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=col_name).font = Font(bold=True)
                    row_idx += 1

                    for _, row in block.iterrows():
                        for col_idx, col_name in enumerate(block.columns, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=row[col_name])
                        row_idx += 1

                    row_idx += 1

            autofit_columns(ws)

            # === Resumo ===
            ws_resumo = workbook.create_sheet(title="Resumo")
            row_idx = 1

            if not df_summary.empty:
                # Drop AAAA columns that are fully empty (any tipo)
                for col in df_summary.columns:
                    if col.endswith("AAAA") and df_summary[col].replace("", pd.NA).isna().all():
                        df_summary.drop(columns=[col], inplace=True)

                tipos_disponiveis = df_summary["Tipo"].dropna().unique().tolist()  # Get unique types from summary data

                for tipo_lower in tipos_disponiveis:
                    tipo_lower = tipo_lower.lower().strip()
                    block = df_summary[df_summary["Tipo"].str.lower().str.strip() == tipo_lower].copy()

                    if block.empty:
                        continue

                    block = block.sort_values(by=["Corretora", "Número da Nota"])

                    # === Get standardized columns from correct RESUMO_KEY_MAP
                    if tipo_lower == "a vista":
                        base_columns = list(RESUMO_KEY_MAP_AVISTA.keys())
                    elif tipo_lower == "bm&f":
                        base_columns = list(RESUMO_KEY_MAP_BMF.keys())
                    else:
                        base_columns = block.columns.tolist()

                    # Interleave AAAA columns only if they exist
                    block_columns = ["Corretora", "Número da Nota"]
                    for col in base_columns:
                        if col in block.columns:
                            block_columns.append(col)
                            aaaa_col = f"{col} AAAA"
                            if aaaa_col in block.columns:
                                block_columns.append(aaaa_col)

                    block = block[[col for col in block_columns if col in block.columns]]

                    ws_resumo.cell(row=row_idx, column=1, value=f"*** {tipo_lower.upper()} ***").font = Font(bold=True)
                    row_idx += 2

                    # Labels (skip header for AAAA columns)
                    for col_idx, col_name in enumerate(block.columns, start=1):
                        label = "" if "AAAA" in col_name else col_name
                        ws_resumo.cell(row=row_idx, column=col_idx, value=label).font = Font(bold=True)

                    row_idx += 1

                    # Data rows
                    for _, row in block.iterrows():
                        for col_idx, col_name in enumerate(block.columns, start=1):
                            ws_resumo.cell(row=row_idx, column=col_idx, value=row[col_name])
                        row_idx += 1

                    row_idx += 1  # Space between sections

            autofit_columns(ws_resumo)

            # === Consistência ===
            if not df_consistency.empty:
                ws_consistencia = workbook.create_sheet(title="Consistência")
                row_idx = 1

                if "Tipo" in df_consistency.columns:
                    tipos_consistencia = df_consistency["Tipo"].dropna().unique().tolist()
                else:
                    tipos_consistencia = ["Unknown"]

                for tipo in tipos_consistencia:
                    tipo_lower = tipo.strip().lower()
                    block = df_consistency[df_consistency["Tipo"].str.lower().str.strip() == tipo_lower].copy()

                    if block.empty:
                        continue

                    block = block.sort_values(by=["Corretora", "Número da Nota"])
                    block_columns = ["Corretora", "Número da Nota", "Tipo"]

                    # === Usar coluna ajustada para negócios individuais
                    temp = df_trades[df_trades["Tipo"].str.lower().str.strip() == tipo_lower]

                    if tipo_lower == "bm&f":
                        if "valor_trades_consistency" in temp.columns:
                            grouped = (
                                temp.groupby(["Número da Nota", "Corretora"])["valor_trades_consistency"]
                                .apply(lambda x: abs(x.sum()))
                                .reset_index(name="Negócios Individuais")
                            )
                            block = block.merge(grouped, on=["Número da Nota", "Corretora"], how="left")
                            block_columns.append("Negócios Individuais")

                    elif tipo_lower == "a vista":
                        if "Valor Operação / Ajuste" in temp.columns:
                            grouped = temp.groupby(["Número da Nota", "Corretora"])["Valor Operação / Ajuste"].sum().reset_index()
                            grouped.rename(columns={"Valor Operação / Ajuste": "Negócios Individuais"}, inplace=True)
                            block = block.merge(grouped, on=["Número da Nota", "Corretora"], how="left")
                            block_columns.append("Negócios Individuais")

                    if "Negócios Individuais" in block.columns and "valor_das_operacoes" in block.columns:
                        block["Diferença"] = (
                            block["Negócios Individuais"] - block["valor_das_operacoes"]
                        ).round(2)
                        block["Status"] = block["Diferença"].apply(
                            lambda x: "OK" if abs(x) < 0.01 else "Inconsistência"
                        )

                    block_columns += ["valor_das_operacoes", "Diferença", "Status"]
                    block = block[[col for col in block_columns if col in block.columns]]

                    # === Título da seção
                    ws_consistencia.cell(row=row_idx, column=1, value=f"*** {tipo.upper()} ***").font = Font(bold=True)
                    row_idx += 2

                    # === Cabeçalhos com formatação
                    for col_idx, col_name in enumerate(block.columns, start=1):
                        new_label = col_name

                        if col_name == "Negócios Individuais":
                            if tipo_lower == "bm&f":
                                new_label = "Valor Operação\n (Negócios Individuais)"
                            elif tipo_lower == "a vista":
                                new_label = "Valor Operação / Ajuste\n (Negócios Individuais)"

                        elif col_name == "valor_das_operacoes":
                            new_label = "Valor das Operações\n (Resumo da Nota)"

                        cell = ws_consistencia.cell(row=row_idx, column=col_idx, value=new_label)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(wrap_text=True, horizontal="center")

                    row_idx += 1

                    # === Linhas de dados
                    for _, row in block.iterrows():
                        for col_idx, col_name in enumerate(block.columns, start=1):
                            ws_consistencia.cell(row=row_idx, column=col_idx, value=row[col_name])
                        row_idx += 1

                    row_idx += 1  # Espaçamento entre seções

                autofit_columns(ws_consistencia)

        pd.DataFrame().to_excel(writer, sheet_name="dummy", index=False)


        return df_trades
